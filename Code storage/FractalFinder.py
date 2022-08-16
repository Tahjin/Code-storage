from cv2 import sqrt
from matplotlib import pyplot as plt
# import numpy as np
import math
from TrigPy import TrigPy
import random
trig=TrigPy
import openpyxl
from openpyxl.utils import get_column_letter
wbook = openpyxl.Workbook()

ws = wbook.active
ws.title = 'Shape_sheet'
ws.append(['xy angle','yz angle','distance'])

z=[]
x=[]
y=[]
random.seed(5) 
for i in range(50):
    
    x.append(random.randint(1, 35))
random.seed(4) 
for i in range(50):
    
    y.append(random.randint(1, 35))
random.seed(10) 
for i in range(50):
    
    z.append(random.randint(1, 35))

print("\n")
for i in range(len(x)):
    if i==len(x)-1:
        break
    else:
        run = x[i+1]-x[i]
        rise = y[i+1]-y[i]
        runY = y[i+1]-y[i]
        riseZ = z[i+1]-z[i]
        distance=round(math.sqrt(pow(x[i+1]-x[i],2)+pow(y[i+1]-y[i],2)+pow(z[i+1]-z[i],2)))
        # print("Point: "+str(x[i])+","+str(y[i])+","+str(z[i]))
        # print("Point: "+str(x[i+1])+","+str(y[i+1])+","+str(z[i+1]))
        # print("x,y angle: "+str(trig.calc_angle(rise,run)+" y,z angle: "+str(trig.calc_angle(riseZ,runY))))
        # print("Distance: "+str(distance))
        # print("\n")
        ws.append([trig.calc_angle(rise,run),trig.calc_angle(riseZ,runY),distance]) 
# print(trig.calc_angle(-3,7))
#
#note now I just need to confirm z angle
#


#remember to close xcel file first before saveing
print(wbook.sheetnames)
wbook.save('C:/Users/Talyn/Documents/Exel sheets1/Fractal1.xlsx')

