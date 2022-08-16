import openpyxl
wbook = openpyxl.load_workbook('C:/Users/Talyn/Documents/Exel sheets1/Lotto_Data.xlsx')
ws = wbook.active


#this will make changes in xcel file
#ws['A5'].value = 'norman'#i only need to do this once

#remember to close xcel file first
column_b=ws['B']
# wbook.save('C:/Users/Talyn/Documents/Exel sheets1/Testxl.xlsx')
bigestNum=0
for cell in column_b:
    templist=cell.value.split()
    if cell.value=="Nunbers":
        pass
    else:
        for num in templist:
            if int(num)>bigestNum:
                bigestNum=int(num)

# templist=column_b[1].value.split()
print(bigestNum)
    # print(cell.value)
# print(column_b[0][0])